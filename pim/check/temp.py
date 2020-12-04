from openpyxl import load_workbook
import time

start = time.time()
wb = load_workbook('./files/TM_CommonToRoss.xlsx')
ws = wb['0 运动-运动鞋-羽毛球鞋']

from openpyxl.cell.cell import Cell
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet import worksheet


# cell = Cell(ws, 1, 1)
# print(cell.value)

# print(ws.cell(3, 1).value)
# print(ws.cell(3, 1).value)

# print(ws['A4'].coordinate)


# print(ws[f'A{4 - 1}'].value)  # nonono
# print(type(ws.merged_cells))
# for c in ws.merged_cells:
#     print(c,type(c))


def get_merged_cell_value(ws: worksheet, cell: Cell):
    '''
    获取 合并的单元格数据，如果单元格没值，则递归向上一个单元格取值
    '''
    if isinstance(cell, MergedCell) and cell.value is None:
        return get_merged_cell_value(ws, ws.cell(cell.row - 1, cell.column))
    return cell.value


def get_sheet_values(ws: worksheet, min_row=None, max_row=None, min_col=None, max_col=None):
    '''
    :param ws: worksheet
    :param min_row:
    :param max_row:
    :param min_col:
    :param max_col:
    :return: generator
    '''

    for row in ws.iter_rows(min_row, max_row, min_col, max_col, values_only=False):
        # cells = (ws.cell(row=row, column=column) for column in range(min_col, max_col + 1))
        v = [get_merged_cell_value(ws, cell) for cell in row]
        yield v


import json

# print(list(ws.values)[1])
g = get_sheet_values(ws, 3)

print(json.dumps(list(g), ensure_ascii=False))

# for row in ws.iter_rows(min_row=1, max_row=5):
#
#     for cell in row:
#         # print(cell.value)
#         # print(f'{cell.coordinate},{cell.row, cell.column},{type(cell)}')
#         print(f'当前坐标：{cell.coordinate},{cell.row, cell.column},{cell.value}')
#         if cell.value is None:
#             v = get_merged_cell_value(ws, cell)
#             print(f'获取到的值是：{v}')
#
#     print('-' * 50)


import json

# print(json.dumps(list(ws.values), ensure_ascii=False))

end = time.time()
print(end - start)
