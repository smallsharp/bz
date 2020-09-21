# coding=utf8
import requests
from openpyxl import load_workbook
from common.tool import GeneralTool


def request(url, data: dict, headers):
    response = requests.post(url, json=data, headers=headers)
    return response


def get_schemalist_from_excel(xlsxfile: str, sheetname):
    '''
    从转换规则excel中，取出schema_code转换规则数据，备用
    :param xlsxfile:
    :param sheetname:
    :return:
    '''
    wb = load_workbook(xlsxfile)

    # print(wb.sheetnames)
    sheet = wb[sheetname]
    rows = list(sheet.rows)
    cols = list(sheet.columns)
    values = list(sheet.values)

    schemalist = list()

    # 跳过第一行
    for v in values[1:]:
        scobj = dict()
        # 序列解包
        deptname, _, _, key1, key2, key3, key4, _, schema_code = v
        # print(deptname, key1, key2, schema_code)
        scobj['deptname'] = deptname
        scobj['schema_code'] = schema_code
        scobj['keyword'] = [key1, key2, key3, key4]
        schemalist.append(scobj)
    return schemalist


def get_schemacode_by_keyword(deptname, keyword: str, schemalist):
    for item in schemalist:
        if deptname == item['deptname'] and (
                item[keyword][0] in keyword or item[keyword][1] in keyword or item[keyword][2] in keyword or
                item[keyword][3] in keyword):
            return item['schema_code']

    return None


def main():
    url = "http://mid-platform-publish-platform-service-sit.cloud.bz/publish-platform/convertJs/hub-convert"
    headers = {
        'tenantCode': 'baozun',
        'catalog': 'anf',
        'ruleCode': 'source-common-ins-upd',
        'Content-Type': 'application/json'
    }

    # 1. 原始数据(ins.csv)，从中 取出deptname,keyword ，并将原始数据转成dict(用来调用hub-convert接口)
    # 2. 通过deptname,keyword 到ANF类目转化中，查到预期的schemaCode_exp
    # 3. 使用dict调用hub-convert接口，查询通过js转换后的schemaCode_act
    # 4. 对比 查到预期的schemaCode_exp 和 通过js转换后的schemaCode_act
    dict_reader = GeneralTool.get_csv_data('../files/ins.csv')
    for data in dict_reader:
        dept_name = data['DEPT_NAME']
        unique_name_cn = data['UNIQUE_NAME_CN']
        print(f'源数据中DEPT_NAME:{dept_name}, UNIQUE_NAME_CN:{unique_name_cn}')
        # print(data)
        # print(json.dumps(data, ensure_ascii=False))
        response = request(url, data, headers=headers)
        variant = response.json()['data']['variants'][0]
        ##################
        # {'variantCode': '952810001', 'schemaCode': 's1', 'properties': {'customColor': {'customColorID': 'C1', 'customColorName': '白色001'}, 'color': '43', 'KIC_ID': 'KIC_322-2110-0986-101', 'COLOR_CD': 'C1', 'COMPLIANCE_TAG_URL': 'http://anf.scene7.com/is/image/anf/sku639861338-cn-comply-label1', 'WEB_LONG_SKU': 'w1', 'customReferenceHeight': 'Slim1', 'customSize': 'L1Slim1', 'PRICE_TICKET_URL': 'http://anf.scene7.com/is/image/anf/sku639861338-price-ticket-RMB1', 'skuCode': 'AFUPC952810001', 'platformCode': 'gn1'}}
        #############
        # 从接口返回值中，查到实际接口返回的schemacode
        schemaCode_act = variant['schemaCode']

        # 通过原始数据中的 deptname 和 unique_name_cn ,从转换规则中，查到预期的schemacode
        schemalist = get_schemalist_from_excel('ANF上新类目转化逻辑-0528.xlsx', 'Sheet1')
        schemaCode_exp = get_schemacode_by_keyword(dept_name, keyword=unique_name_cn, schemalist=schemalist)

        print(f'{schemaCode_exp == schemaCode_act},schemaCode_exp:{schemaCode_exp},schemaCode_act:{schemaCode_act}')

        break


if __name__ == '__main__':
    # main()
    schemalist = get_schemalist_from_excel('ANF上新类目转化逻辑-0528.xlsx', 'Sheet1')
    print(schemalist)

    print(get_schemacode_by_keyword('ANF WOMENS JEANS', '高腰妈咪牛仔裤', schemalist))
