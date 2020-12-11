import pymongo
import csv
from openpyxl import load_workbook
from openpyxl import Workbook
import json
import os
from openpyxl.cell.cell import Cell
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet import worksheet


class GeneralTool:

    @staticmethod
    def get_collection(uri, dbname, collection, authuser=None, authpwd=None, **kwargs):
        # 1. 连接mongo服务,返回client对象
        try:
            client = pymongo.MongoClient(uri)
        except Exception as e:
            raise Exception("connect mongo error====>>>\n", e)
        # database = client.get_database(dbname)  # mydb = client[dbname]
        database = client[dbname]  # print(database.list_collection_names())

        if authuser and authpwd:
            database.authenticate(authuser, authpwd)

        # 3. use collection(table)
        mycol = database.get_collection(collection)  # mycol = database[collection]

        # 4. close
        client.close()

        return mycol

    @classmethod
    def get_mongo_data(self, mongouri, mongodb, mongocol, query_expression: dict):
        self.mycol = self.get_collection(mongouri, mongodb, mongocol)
        return self.mycol.find(query_expression)

    @staticmethod
    def get_csv_data(csvfile: str, delimiter='|', skip_first_row=False, encoding='utf8'):
        csvfile = open(csvfile, 'r', encoding=encoding)
        # fieldnames = ('upc', 'brand', 'name', 'enName', 'listPrice', 'colorCd', 'colorName')
        # reader = csv.DictReader(csvfile, fieldnames=fieldnames, delimiter='|')
        reader = csv.DictReader(csvfile, delimiter=delimiter)
        # 跳过第一行，如果指定了上面的fieldnames, 需要跳过第一行title数据
        if skip_first_row:
            next(reader)
        return reader

    @classmethod
    def compare(self, datas_mongo, datas_csv):
        print('compare')
        print(
            "mongo数据条数：{},csv数据条数：{}".format(len(datas_mongo := list(datas_mongo)), len(datas_csv := list(datas_csv))))
        if not datas_mongo or len(list(datas_mongo)) == 0:
            raise Exception('mongo没有检测到数据')

        if not datas_csv or len(list(datas_csv)) == 0:
            raise Exception('csv没有检测到数据')

        if len(list(datas_mongo)) != len(list(datas_csv)):
            raise Exception(
                'csv数据数量:{},mongo数据数量{},数量不一致，无法比较！'.format(len(list(datas_mongo)), len(list(datas_csv))))

        for m, c in zip(datas_mongo, datas_csv):
            # print(m['sourceData'])
            # print(m==c)
            s = m['sourceData']
            # print(json.dumps(s,ensure_ascii=False))
            # print(json.dumps(c,ensure_ascii=False))
            # print(f'mongo.barCode:{s["BARCODE"]},csv.barCode:{c["BARCODE"]}')
            # 查看两个dict 不共有的key
            if s.keys() ^ c.keys():
                print('不同的key：', s.keys() ^ c.keys())
                print('前有后无的key：', s.keys() - c.keys())
                print('前有后无的key：', c.keys() - s.keys())

            # 查看两个dict 共有的key
            # print(s.keys() & c.keys())

            # same key  diff value
            samekey = s.keys() & c.keys()
            # print(samekey)
            diff_vals = [{k: (s[k], c[k])} for k in samekey if s[k] != c[k]]
            print(f'异常的数据：{json.dumps(diff_vals, ensure_ascii=False)}')
            print('------' * 40)


class ExcelTool:

    def __init__(self, filepath=None):
        self.filepath = filepath
        self.__load_workbook()

    def __load_workbook(self):
        if self.filepath and os.path.exists(self.filepath):
            self.workbook = load_workbook(self.filepath)
        else:
            self.workbook = Workbook()

    def get_worksheet(self, name):
        return self.workbook[name]

    def get_active_worksheet(self):
        return self.workbook.active

    def get_sheetnames(self):
        return self.workbook.sheetnames

    def get_merged_cell_value(self, ws: worksheet, cell: Cell):
        '''
        获取 合并的单元格数据，如果单元格没值，则递归向上一个单元格取值
        '''
        if isinstance(cell, MergedCell) and cell.value is None:
            return self.get_merged_cell_value(ws, ws.cell(cell.row - 1, cell.column))
        return cell.value

    def get_sheet_values(self, sheetname, skiplines=1):
        '''
        [('9759', 's1011713', None, 'customSize'), ('9759', 's1011713', 'customcolor.customcolorname', None)]
        :param sheetname:
        :param skiplines: lines to skip, like skip 2 line;
        :return: 如上格式 行列表
        '''
        ws = self.workbook.get_sheet_by_name(sheetname)

        values = list(ws.values)  # generator > list
        if skiplines > len(values):
            raise ValueError('there is no more lines to skip!')
        return values[skiplines:]

    def get_sheet_values2(self, ws: worksheet, min_row=None, max_row=None, min_col=None, max_col=None):
        '''
        :param ws: worksheet
        :param min_row:
        :param max_row:
        :param min_col:
        :param max_col:
        :return: generator
        '''
        for row in ws.iter_rows(min_row, max_row, min_col, max_col, values_only=False):
            # cells = (ws.cell(row=row, column=column) for column in range(min_col, max_col + 1))
            v = [self.get_merged_cell_value(ws, cell) for cell in row]
            yield v

    def merge_dict(self, sheetname, dictname='dictname', merge_column='1', value_column=2, skiplines=1):
        '''
        读取sheet 并将数据 封装成dict
        :param sheetname: 数据所在的sheetname,like 'sheet1'
        :param dictname: 给新转换后的数据 起个名字, like 'title'
        :param skiplines: 需要跳过的行数，默认跳过第一行
        :return: 返回第一字典,格式如下
        {
            "name": "xxx",
            "maps": {
                "CLASSIC_BAG": "运动包/户外包/配件 > 挎包/拎包/休闲包s1011105",
                "TRAINING_BAG": "运动包/户外包/配件 > 挎包/拎包/休闲包s1011105",
                "TRAINING_OTHERS": "运动包/户外包/配件 > 其他服饰配件s1011107",
                "CLASSIC_GLOVES": "运动包/户外包/配件>手套s1011111"
            }
        }
        '''
        maps = dict()
        for row in self.get_sheet_values(sheetname, skiplines):
            # 值 取对应列的值，excel列从1开始，row（元组）是从0开始，需要减去1
            value = row[int(value_column) - 1]
            if len(merge_column.split(',')) == 1:
                key = row[int(merge_column) - 1]
                maps[str(key)] = value
            else:
                # 如果按","分隔后有多条,则需要分别取出列 对应的值，放入keys中，然后使用"_"连接拼成一个key
                column_list = merge_column.split(',')
                keys = list()
                for column in column_list:
                    key_value = str(row[int(column) - 1]).strip()  # 去掉两边的空格
                    keys.append(key_value)
                key = "_".join(keys)
                maps[str(key)] = value

        return {'name': dictname, 'maps': maps}

    def get_schemalist_from_sheet(self, sheetname, skiplines=1):
        '''
        获取 sheet中的类目信息
        :param sheetname:
        :param skiplines:1 默认跳过第一行
        :return:
        '''
        schemalist = list()
        last_schemacode = None

        for row in self.get_sheet_values(sheetname, skiplines):
            keyword1, keyword2, _, schemacode = row
            if schemacode:
                last_schemacode = schemacode
            else:
                schemacode = last_schemacode
            # 如果是OTHER 这个sheet，需要拼接关键字返回
            obj = dict()

            if sheetname == 'OTHER':
                # keyword1 = f'{keyword1} {keyword2}'
                # 非鞋类，使用如下关键字
                obj['category'] = keyword1
                obj['local_product_type'] = keyword2
            else:
                obj['keyword'] = keyword1
            obj['schemacode'] = schemacode
            schemalist.append(obj)
        return schemalist

    @classmethod
    def write(cls, datas: list, title, startrow=1, startcol=1, filename='myexcel.xlsx',
              sheetname='report'):
        '''
        :param datas: 需要写入的内容【二维数据】,like [(111, 222, 333, 444), ('aaa', 'bbb', 'ccc'), ('aaa1', 'bbb2', 'ccc3', 'ddd4', 'eee5')]
        :param title: 需要写入的标题 【一维数组】
        :param filename: 生成的excel名称 like myexcel.xlsx
        :param sheetname: 生成的sheet名称 like sheet1
        :return: generate xx.xlsx file
        '''
        workbook = Workbook()
        # sheet = workbook.create_sheet(title="report")
        sheet = workbook.active
        sheet.title = sheetname  # sheet = workbook.create_sheet(title="report")

        if title and isinstance(title, list):
            for col, v in enumerate(title):
                sheet.cell(startrow, col + startcol, title[col])
            startrow += 1
        for row, item in enumerate(datas, start=0):  # 索引从1 开始
            for col, value in enumerate(item, start=0):
                # 默认从第一行，第一列开始写 (row, column, value=None)
                sheet.cell(row + startrow, col + startcol, str(datas[row][col]))

        for col, value in enumerate(datas, start=0):
            # 默认从第一行，第一列开始写 (row, column, value=None)
            sheet.cell(startrow + 1, col + startcol, str(value))
            sheet.append(datas)
            # save excel
        workbook.save(filename=filename)
        workbook.close()

    def append(self, sheetname, datas: list):
        '''
        u need save in another method
        :param filename:
        :param sheetname:
        :param datas: 一维数组
        :return:
        '''
        if sheetname not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheetname, index=0)

        ws = self.workbook[sheetname]
        ws.append([str(item) for item in datas])

    def save(self):
        self.workbook.save(self.filepath)

    def merge_row(self, sheetname, idx_keyrow=0, idx_valuerow=1, skip_lines=0):
        '''
        :param sheetname:
        :param idx_keyrow: 第几行作为key 行
        :param skip_lines:
        :return: key:value 字典数据
        '''
        values = self.get_sheet_values(sheetname, skiplines=skip_lines)
        return dict(zip(values[idx_keyrow], values[idx_valuerow]))

    def merge_list(self, valuelist, idx_key_row=0, idx_value_row=1):
        '''
        :param sheetname:
        :param idx_key_row: 第几行作为key 行
        :param skip_lines:
        :return: key:value 字典数据
        '''
        return dict(zip(valuelist[idx_key_row], valuelist[idx_value_row]))


class CommonTool():
    @classmethod
    def get_value_from_dict(cls, dictname, keyword, rulelist):
        for item in rulelist:
            if item['name'] != dictname:
                continue
            for key in (maps := item['maps']):
                if key.upper() == keyword.upper():
                    return maps.get(key)

        return None

    @classmethod
    def get_maps(cls, dictname, rulelist):
        for item in rulelist:
            if item['name'] == dictname:
                return item['maps']
        return None

    @classmethod
    def get_matchvalue_from_dict(cls, dictname, keyword: str, rulelist, ignoreCase=True):
        maps = cls.get_maps(dictname, rulelist)
        # print(maps)
        if not maps:
            return
        for key in maps.keys():
            if ignoreCase:
                keyword = keyword.upper()
                # key = keyword.upper()
                if keyword.find(key.upper()) != -1:
                    return maps[key]
            else:
                if keyword.find(key) != -1:
                    return maps[key]


if __name__ == '__main__':
    # e = ExcelTool('xx.xlsx')
    # print(e.get_sheet_values('report'))
    #
    # print(e.get_dict_rows('report', 0, 1))

    # title = ["t1", "t2", "t3"]
    #
    # # reportlist = ['a', 'b', ['c1', 'c2']]
    # reportlist = [['a', 'b', 'c'], ['a', 'b', ["xx1", "xx2"]]]
    # e.write(reportlist, title, filename='xx.xlsx')

    items = [
        {
            "name": "customColor.customColorName",
            "maps": {
                "白色": {
                    "属性键值": "customColor.customColorName",
                    "属性名称": "白色",
                    "属性值名称": "白色",
                    "属性值键值": "白色",
                    "京东字段": "163785",
                    "京东字段名称": "颜色",
                    "京东属性值key": "白色",
                    "京东属性值value": "835637"
                },
                "黑色": {
                    "属性键值": "customColor.customColorName",
                    "属性名称": "黑色",
                    "属性值名称": "黑色",
                    "属性值键值": "黑色",
                    "京东字段": "163785",
                    "京东字段名称": "颜色",
                    "京东属性值key": "黑色",
                    "京东属性值value": "835644"
                }
            }
        },
        {
            "name": "sneakerTechnology",
            "maps": {
                "Flexiblefunction": {
                    "属性键值": "sneakerTechnology",
                    "属性名称": "运动鞋科技",
                    "属性值名称": "易弯折功能",
                    "属性值键值": "Flexiblefunction",
                    "京东字段": "162846",
                    "京东字段名称": "运动鞋科技",
                    "京东属性值key": "易弯折功能",
                    "京东属性值value": "832677"
                },
                "naturalfeet": {
                    "属性键值": "sneakerTechnology",
                    "属性名称": "运动鞋科技",
                    "属性值名称": "天足",
                    "属性值键值": "naturalfeet",
                    "京东字段": "162846",
                    "京东字段名称": "运动鞋科技",
                    "京东属性值key": "天足",
                    "京东属性值value": "832670"
                }
            }
        }
    ]

    print(CommonTool.get_matchvalue_from_dict('sneakerTechnology', 'naturalfeet', items))
